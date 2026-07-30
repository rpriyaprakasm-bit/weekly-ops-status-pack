"""
Collect weekly ops status inputs from Excel (.xlsx) or CSV folder.

Excel path (preferred): data/ops_status_input.xlsx with sheets:
  Header, Headcount, Programs, Projects, ActionItems, Wins

CSV path (SharePoint export / simple edit):
  data/csv/Header.csv, Headcount.csv, Programs.csv, Projects.csv,
  ActionItems.csv, Wins.csv

SharePoint pattern:
  Maintain a list (or Excel in a document library) with the same columns →
  Export to Excel/CSV → place under data/ → run workflow.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "data"
CSV_DIR = DATA / "csv"
XLSX = DATA / "ops_status_input.xlsx"


def _read_csv(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def from_csv_folder() -> dict:
    header_rows = _read_csv(CSV_DIR / "Header.csv")
    header = {r["Field"].strip(): r["Value"].strip() for r in header_rows if r.get("Field")}

    hc_rows = _read_csv(CSV_DIR / "Headcount.csv")
    hc = hc_rows[0] if hc_rows else {}
    headcount = {
        "total": _int(hc.get("total")),
        "billable": _int(hc.get("billable")),
        "non_billable": _int(hc.get("non_billable")),
        "billable_pct": _int(hc.get("billable_pct")),
        "target_billable_pct": _int(hc.get("target_billable_pct")),
        "bench": _int(hc.get("bench")),
        "notes": (hc.get("notes") or "").strip(),
    }

    programs = []
    for r in _read_csv(CSV_DIR / "Programs.csv"):
        programs.append({
            "name": (r.get("name") or "").strip(),
            "owner": (r.get("owner") or "").strip(),
            "rag": (r.get("rag") or "On Track").strip(),
            "next_milestone": (r.get("next_milestone") or "").strip(),
            "update": (r.get("update") or "").strip(),
        })

    projects = []
    for r in _read_csv(CSV_DIR / "Projects.csv"):
        billable = str(r.get("billable", "")).strip().lower() in ("yes", "true", "1", "y")
        projects.append({
            "name": (r.get("name") or "").strip(),
            "program": (r.get("program") or "").strip(),
            "owner": (r.get("owner") or "").strip(),
            "rag": (r.get("rag") or "On Track").strip(),
            "start": (r.get("start") or "").strip(),
            "target_end": (r.get("target_end") or "").strip(),
            "update": (r.get("update") or "").strip(),
            "billable": billable,
        })

    actions = []
    for r in _read_csv(CSV_DIR / "ActionItems.csv"):
        actions.append({
            "id": (r.get("id") or "").strip(),
            "action": (r.get("action") or "").strip(),
            "owner": (r.get("owner") or "").strip(),
            "due": (r.get("due") or "").strip(),
            "priority": (r.get("priority") or "Medium").strip(),
            "status": (r.get("status") or "Open").strip(),
        })

    wins = [(r.get("win") or "").strip() for r in _read_csv(CSV_DIR / "Wins.csv") if (r.get("win") or "").strip()]

    return _assemble(header, headcount, programs, projects, actions, wins)


def from_xlsx(path: Path | None = None) -> dict:
    path = path or XLSX
    if not path.exists():
        raise FileNotFoundError(f"Excel input not found: {path}")

    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("openpyxl is required for Excel input: pip install openpyxl") from e

    wb = load_workbook(path, data_only=True)

    def sheet_rows(name: str) -> list[dict]:
        if name not in wb.sheetnames:
            return []
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        out = []
        for row in rows[1:]:
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            out.append({headers[i]: ("" if row[i] is None else row[i]) for i in range(len(headers))})
        return out

    header = {}
    for r in sheet_rows("Header"):
        field = str(r.get("Field", "")).strip()
        if field:
            header[field] = str(r.get("Value", "")).strip()

    hc_rows = sheet_rows("Headcount")
    hc = hc_rows[0] if hc_rows else {}
    headcount = {
        "total": _int(hc.get("total")),
        "billable": _int(hc.get("billable")),
        "non_billable": _int(hc.get("non_billable")),
        "billable_pct": _int(hc.get("billable_pct")),
        "target_billable_pct": _int(hc.get("target_billable_pct")),
        "bench": _int(hc.get("bench")),
        "notes": str(hc.get("notes") or "").strip(),
    }

    programs = []
    for r in sheet_rows("Programs"):
        programs.append({
            "name": str(r.get("name") or "").strip(),
            "owner": str(r.get("owner") or "").strip(),
            "rag": str(r.get("rag") or "On Track").strip(),
            "next_milestone": str(r.get("next_milestone") or "").strip(),
            "update": str(r.get("update") or "").strip(),
        })

    projects = []
    for r in sheet_rows("Projects"):
        b = str(r.get("billable") or "").strip().lower()
        projects.append({
            "name": str(r.get("name") or "").strip(),
            "program": str(r.get("program") or "").strip(),
            "owner": str(r.get("owner") or "").strip(),
            "rag": str(r.get("rag") or "On Track").strip(),
            "start": str(r.get("start") or "").strip()[:10],
            "target_end": str(r.get("target_end") or "").strip()[:10],
            "update": str(r.get("update") or "").strip(),
            "billable": b in ("yes", "true", "1", "y"),
        })

    actions = []
    for r in sheet_rows("ActionItems"):
        due = str(r.get("due") or "").strip()[:10]
        actions.append({
            "id": str(r.get("id") or "").strip(),
            "action": str(r.get("action") or "").strip(),
            "owner": str(r.get("owner") or "").strip(),
            "due": due,
            "priority": str(r.get("priority") or "Medium").strip(),
            "status": str(r.get("status") or "Open").strip(),
        })

    wins = [str(r.get("win") or "").strip() for r in sheet_rows("Wins") if str(r.get("win") or "").strip()]

    return _assemble(header, headcount, programs, projects, actions, wins)


def collect() -> dict:
    """Prefer Excel workbook; fall back to CSV folder (SharePoint export friendly)."""
    if XLSX.exists():
        print(f"Collecting from Excel: {XLSX}")
        return from_xlsx(XLSX)
    print(f"Collecting from CSV folder: {CSV_DIR}")
    return from_csv_folder()


def write_workstream_json(payload: dict, dest: Path | None = None) -> Path:
    dest = dest or (DATA / "workstream_updates.json")
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {dest}")
    return dest


def _assemble(header, headcount, programs, projects, actions, wins) -> dict:
    at_risk = [p for p in projects if "risk" in (p.get("rag") or "").lower() or "block" in (p.get("rag") or "").lower()]
    on_track = [p for p in projects if "track" in (p.get("rag") or "").lower()]
    return {
        "week_ending": header.get("week_ending", ""),
        "report_date": header.get("report_date", header.get("week_ending", "")),
        "period": header.get("period", ""),
        "program": header.get("program", "Business Operations"),
        "prepared_for": header.get("prepared_for", "Weekly Business Operations Review"),
        "overall_business_health": header.get("overall_business_health", "Watch"),
        "overall_rag": header.get("overall_rag", "At Risk"),
        "headcount": headcount,
        "programs": programs,
        "projects": projects,
        "projects_at_risk": [
            {
                "name": p["name"],
                "owner": p["owner"],
                "rag": p["rag"],
                "target_end": p.get("target_end", ""),
                "summary_line": p.get("update", ""),
                "billable": p.get("billable", False),
            }
            for p in at_risk
        ],
        "action_items": actions,
        "wins": wins,
        "metrics": {
            "active_projects": len(projects),
            "projects_on_track": len(on_track),
            "projects_at_risk": len(at_risk),
            "projects_blocked": len([p for p in projects if "block" in (p.get("rag") or "").lower()]),
            "programs_total": len(programs),
            "programs_on_track": len([p for p in programs if "track" in (p.get("rag") or "").lower()]),
            "programs_at_risk": len([p for p in programs if "risk" in (p.get("rag") or "").lower() or "block" in (p.get("rag") or "").lower()]),
            "open_actions": len([a for a in actions if (a.get("status") or "").lower() not in ("done", "closed")]),
            "high_priority_actions": len([a for a in actions if (a.get("priority") or "").lower() == "high"]),
            "client_projects": len([p for p in projects if p.get("billable")]),
            "internal_projects": len([p for p in projects if not p.get("billable")]),
        },
        "cross_cutting_notes": [],
    }


def _int(v):
    try:
        if v is None or v == "":
            return None
        return int(float(v))
    except (TypeError, ValueError):
        return None


if __name__ == "__main__":
    payload = collect()
    write_workstream_json(payload)
